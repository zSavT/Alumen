# --- START OF FILE AlumenCore.py ---
import time
import google.generativeai as genai
import google.api_core.exceptions
import csv
import os
import re
import argparse
import sys
import json
import polib
import threading
from threading import Lock
from datetime import datetime
from tenacity import retry, stop_after_attempt, wait_exponential
from collections import deque

# Tentativo import moduli opzionali
try:
    import openpyxl
except ImportError:
    openpyxl = None

from rich.console import Console

# ----- CONFIGURAZIONE VERSIONE 2.2 -----
CURRENT_SCRIPT_VERSION = "2.2.0"
DEFAULT_MODEL_NAME = "gemini-2.0-flash"
CACHE_FILE_NAME = "alumen_cache.json"
LOG_FILE_NAME = "log.txt"
ESTIMATED_CHARS_PER_TOKEN = 3.5

# Dizionario Descrizioni Modelli (per la GUI)
MODEL_DESCRIPTIONS = {
    "gemini-2.0-flash": "Ultima versione: Velocissimo, intelligente, multimodale.",
    "gemini-2.0-flash-lite-preview": "Versione leggera della 2.0, ottimizzata per costi.",
    "gemini-1.5-flash": "Stabile: Ottimo bilanciamento velocità/costo per grandi volumi.",
    "gemini-1.5-flash-8b": "Versione ultra-leggera, minor costo in assoluto.",
    "gemini-1.5-pro": "Alta Intelligenza: Più lento e costoso, ma migliore qualità.",
    "gemini-1.0-pro": "Legacy: Prima generazione stabile."
}

# ----- GLOBALI -----
console = Console()
translation_cache = {}
available_api_keys = []
current_api_key_index = 0
model = None
glossary_terms = {}
api_call_counts = {}
total_files_translated = 0
rpm_limit = None
rpm_request_timestamps = []
rpm_lock = Lock()
last_cache_save_time = 0
gui_log_queue = None 

def log_msg(message, style=""):
    timestamp = datetime.now().strftime('%H:%M:%S')
    full_msg = f"[{timestamp}] {message}"
    console.print(message, style=style)
    try:
        with open(LOG_FILE_NAME, 'a', encoding='utf-8') as f:
            f.write(full_msg + "\n")
    except: pass
    if gui_log_queue:
        gui_log_queue.put(full_msg)

def fetch_available_models(api_key):
    """Interroga l'API e restituisce una lista formattata con descrizioni."""
    try:
        genai.configure(api_key=api_key)
        models_list = []
        for m in genai.list_models():
            if 'generateContent' in m.supported_generation_methods:
                name = m.name.replace('models/', '')
                # Aggiungi descrizione se nota
                desc = MODEL_DESCRIPTIONS.get(name, "Modello generico")
                display_name = f"{name}  [{desc}]"
                models_list.append(display_name)
        
        # Ordina mettendo i flash 2.0 e 1.5 in cima
        models_list.sort(key=lambda x: "flash" not in x)
        return models_list
    except Exception as e:
        return [f"Errore: Chiave non valida o problema di rete"]

def setup_engine(args):
    global available_api_keys, model, glossary_terms, translation_cache, rpm_limit

    if args.persistent_cache and os.path.exists(CACHE_FILE_NAME):
        try:
            with open(CACHE_FILE_NAME, 'r', encoding='utf-8') as f:
                translation_cache = json.load(f)
            log_msg(f"💾 Cache caricata: {len(translation_cache)} voci.", style="dim")
        except: pass

    if args.glossary and os.path.exists(args.glossary):
        try:
            with open(args.glossary, 'r', encoding='utf-8') as f:
                for row in csv.reader(f):
                    if len(row) >= 2: glossary_terms[row[0].strip()] = row[1].strip()
            log_msg(f"📚 Glossario caricato: {len(glossary_terms)} termini.", style="green")
        except Exception as e:
            log_msg(f"⚠️ Errore Glossario: {e}", style="red")

    keys = []
    if args.api:
        keys.extend([k.strip() for k in args.api.split(',') if k.strip()])
    if os.path.exists("api_key.txt"):
        with open("api_key.txt", "r") as f:
            keys.extend([line.strip() for line in f if line.strip()])
    
    available_api_keys = list(dict.fromkeys(keys))
    if not available_api_keys:
        log_msg("🛑 ERRORE: Nessuna API Key trovata.", style="bold red")
        return False

    sys_instr = f"Sei un traduttore esperto da {args.source_lang} a {args.target_lang}."
    if glossary_terms:
        sys_instr += f"\nUsa TASSATIVAMENTE questo glossario:\n{json.dumps(glossary_terms, ensure_ascii=False)}"

    try:
        genai.configure(api_key=available_api_keys[0])
        # Pulisci il nome del modello dalle descrizioni della GUI (es. "gemini-1.5 [Desc]" -> "gemini-1.5")
        clean_model_name = args.model_name.split(' [')[0].strip()
        
        model = genai.GenerativeModel(clean_model_name, system_instruction=sys_instr)
        log_msg(f"✅ Motore Alumen Inizializzato ({clean_model_name})", style="bold cyan")
    except Exception as e:
        log_msg(f"🛑 Errore Init AI: {e}", style="bold red")
        return False
    
    if args.rpm: rpm_limit = args.rpm
    return True

def check_and_save_cache(args, force=False):
    global last_cache_save_time
    if not args.persistent_cache: return
    now = time.time()
    if force or (now - last_cache_save_time > 300):
        try:
            with open(CACHE_FILE_NAME, 'w', encoding='utf-8') as f:
                json.dump(translation_cache, f, ensure_ascii=False, indent=4)
            last_cache_save_time = now
        except: pass

def rotate_key(args):
    global current_api_key_index, model
    current_api_key_index = (current_api_key_index + 1) % len(available_api_keys)
    new_key = available_api_keys[current_api_key_index]
    log_msg(f"🔄 Rotazione API Key -> ...{new_key[-4:]}", style="yellow")
    try:
        genai.configure(api_key=new_key)
        sys_instr = model._system_instruction if hasattr(model, '_system_instruction') else None
        # Estrai clean name anche qui
        clean_model_name = args.model_name.split(' [')[0].strip()
        model = genai.GenerativeModel(clean_model_name, system_instruction=sys_instr)
    except: pass

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
def call_ai_raw(prompt, args):
    if rpm_limit:
        while True:
            with rpm_lock:
                now = time.time()
                global rpm_request_timestamps
                rpm_request_timestamps = [t for t in rpm_request_timestamps if t > now - 60]
                if len(rpm_request_timestamps) < rpm_limit:
                    rpm_request_timestamps.append(now)
                    break
            time.sleep(1)
    time.sleep(0.5)
    try:
        return model.generate_content(prompt).text.strip()
    except Exception as e:
        if args.rotate_on_limit_or_error:
            rotate_key(args)
            raise e
        raise e

def translate_batch(entries, args, stop_event):
    batches = []
    current_batch = []
    current_tokens = 0
    limit_tokens = 3000
    
    for entry in entries:
        text = entry['text']
        cache_key = json.dumps((text, args.source_lang, args.target_lang), ensure_ascii=False)
        
        if cache_key in translation_cache:
            entry['callback'](translation_cache[cache_key])
            continue

        toks = len(text) // ESTIMATED_CHARS_PER_TOKEN
        if (len(current_batch) >= args.batch_size) or (current_tokens + toks > limit_tokens):
            batches.append(current_batch)
            current_batch = []
            current_tokens = 0
        current_batch.append(entry)
        current_tokens += toks
    if current_batch: batches.append(current_batch)

    total_b = len(batches)
    for i, batch in enumerate(batches):
        if stop_event and stop_event.is_set(): return
        texts = [x['text'] for x in batch]
        log_msg(f"    ☁️  Batch {i+1}/{total_b} ({len(texts)} frasi)...", style="dim")

        prompt = f"""TRADUZIONE JSON ARRAY.
Da: {args.source_lang} | A: {args.target_lang}
Rispondi SOLO con un Array JSON di stringhe. Mantieni l'ordine.
Input:
{json.dumps(texts, ensure_ascii=False)}"""

        try:
            resp = call_ai_raw(prompt, args)
            clean = re.sub(r'^```json\s*|\s*```$', '', resp, flags=re.MULTILINE)
            trads = json.loads(clean)
            if len(trads) != len(texts): raise ValueError("Length mismatch")
            
            for idx, t in enumerate(trads):
                ck = json.dumps((texts[idx], args.source_lang, args.target_lang), ensure_ascii=False)
                translation_cache[ck] = t
                batch[idx]['callback'](t)
            check_and_save_cache(args)
                
        except Exception as e:
            log_msg(f"⚠️ Batch fallito ({e}). Riprovo singolarmente.", style="yellow")
            for item in batch:
                try:
                    r = call_ai_raw(f"Traduci in {args.target_lang}: {item['text']}", args)
                    item['callback'](r)
                except: pass

def process_csv(fpath, outpath, args, stop_event):
    with open(fpath, 'r', encoding='utf-8', newline='') as f:
        rows = list(csv.reader(f, delimiter=args.delimiter))
    entries = []
    out_rows = [r[:] for r in rows]
    for i, row in enumerate(out_rows):
        if i > 0 and len(row) > args.translate_col:
            txt = row[args.translate_col]
            if txt.strip() and not txt.isdigit():
                def cb(t, r=row, c=args.output_col):
                    while len(r) <= c: r.append('')
                    r[c] = t
                entries.append({'text': txt, 'callback': cb})
    translate_batch(entries, args, stop_event)
    with open(outpath, 'w', encoding='utf-8', newline='') as f:
        csv.writer(f, delimiter=args.delimiter).writerows(out_rows)

def process_json(fpath, outpath, args, stop_event):
    with open(fpath, 'r', encoding='utf-8') as f: data = json.load(f)
    entries = []
    keys = set(args.json_keys.split(',')) if args.json_keys else set()
    def traverse(obj, path=""):
        if isinstance(obj, dict):
            for k, v in obj.items():
                curr = f"{path}.{k}" if path else k
                match = curr in keys if args.match_full_json_path else k in keys
                if match and isinstance(v, str):
                    entries.append({'text': v, 'callback': lambda t, o=obj, key=k: o.__setitem__(key, t)})
                traverse(v, curr)
        elif isinstance(obj, list):
            for i, x in enumerate(obj): traverse(x, f"{path}[{i}]")
    traverse(data)
    translate_batch(entries, args, stop_event)
    with open(outpath, 'w', encoding='utf-8') as f: json.dump(data, f, indent=4, ensure_ascii=False)

def process_po(fpath, outpath, args, stop_event):
    po = polib.pofile(fpath, encoding='utf-8')
    entries = []
    for entry in po:
        if entry.msgid and not entry.msgid.isdigit():
            def cb(t, e=entry): e.msgstr = t
            entries.append({'text': entry.msgid, 'callback': cb})
    translate_batch(entries, args, stop_event)
    po.save(outpath)

def process_xlsx(fpath, outpath, args, stop_event):
    if not openpyxl:
        log_msg("❌ Modulo 'openpyxl' mancante.", style="red")
        return
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    entries = []
    for row in ws.iter_rows():
        cell = row[0]
        if cell.value and isinstance(cell.value, str):
            tgt_cell = ws.cell(row=cell.row, column=2)
            entries.append({'text': cell.value, 'callback': lambda t, c=tgt_cell: setattr(c, 'value', t)})
    translate_batch(entries, args, stop_event)
    wb.save(outpath)

def run_core_process(args, log_queue=None, stop_event=None):
    global gui_log_queue, total_files_translated
    gui_log_queue = log_queue
    if not setup_engine(args): return
    files = [os.path.join(r, f) for r, _, fs in os.walk(args.input) for f in fs if f.lower().endswith(f".{args.file_type}")]
    if not files:
        log_msg(f"❌ Nessun file .{args.file_type} trovato.", style="red")
        return
    base_out = f"{args.input}_tradotto"
    log_msg(f"🚀 Avvio Alumen su {len(files)} file...", style="bold green")
    for i, fpath in enumerate(files):
        if stop_event and stop_event.is_set():
            log_msg("🛑 Stop utente.", style="red")
            break
        fname = os.path.basename(fpath)
        log_msg(f"📄 [{i+1}/{len(files)}] {fname}")
        rel = os.path.relpath(fpath, args.input)
        out = os.path.join(base_out, rel)
        os.makedirs(os.path.dirname(out), exist_ok=True)
        try:
            if args.file_type == 'csv': process_csv(fpath, out, args, stop_event)
            elif args.file_type == 'json': process_json(fpath, out, args, stop_event)
            elif args.file_type == 'po': process_po(fpath, out, args, stop_event)
            elif args.file_type == 'xlsx': process_xlsx(fpath, out, args, stop_event)
            total_files_translated += 1
            check_and_save_cache(args, force=True)
        except Exception as e:
            log_msg(f"❌ Errore su {fname}: {e}", style="bold red")
    log_msg(f"✅ Completato! {total_files_translated} file tradotti.", style="bold green")

def get_cli_args():
    parser = argparse.ArgumentParser(description="Alumen 2.2 CLI")
    parser.add_argument("--input", default="input")
    parser.add_argument("--api", help="API Key")
    parser.add_argument("--model-name", help="Modello Gemini")
    parser.add_argument("--file-type", default="csv")
    parser.add_argument("--batch-size", type=int, default=30)
    parser.add_argument("--source-lang", default="inglese")
    parser.add_argument("--target-lang", default="italiano")
    parser.add_argument("--glossary")
    parser.add_argument("--persistent-cache", action="store_true")
    parser.add_argument("--rotate-on-limit-or-error", action="store_true")
    parser.add_argument("--rpm", type=int)
    parser.add_argument("--json-keys")
    parser.add_argument("--match-full-json-path", action="store_true")
    args = parser.parse_args()
    if not args.model_name: args.model_name = DEFAULT_MODEL_NAME
    args.delimiter = ","
    args.translate_col = 3
    args.output_col = 3
    return args

if __name__ == "__main__":
    args = get_cli_args()
    run_core_process(args)        model = genai.GenerativeModel(args.model_name, system_instruction=system_instruction)
        log_msg(f"✅ Motore Alumen 2.0 Inizializzato ({args.model_name})", style="bold cyan")
    except Exception as e:
        log_msg(f"🛑 Errore Init AI: {e}", style="bold red")
        return False
    
    if args.rpm: rpm_limit = args.rpm
    return True

def save_cache(args):
    if args.persistent_cache:
        try:
            with open(CACHE_FILE_NAME, 'w', encoding='utf-8') as f:
                json.dump(translation_cache, f, ensure_ascii=False, indent=4)
        except: pass

# ----- LOGICA TRADUZIONE -----

def rotate_key(args):
    global current_api_key_index, model
    current_api_key_index = (current_api_key_index + 1) % len(available_api_keys)
    new_key = available_api_keys[current_api_key_index]
    log_msg(f"🔄 Rotazione API Key -> ...{new_key[-4:]}", style="yellow")
    try:
        genai.configure(api_key=new_key)
        # Ricrea modello
        sys_instr = model._system_instruction if hasattr(model, '_system_instruction') else None
        model = genai.GenerativeModel(args.model_name, system_instruction=sys_instr)
    except: pass

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
def call_ai_raw(prompt, args):
    # RPM Limiter
    if rpm_limit:
        while True:
            with rpm_lock:
                now = time.time()
                global rpm_request_timestamps
                rpm_request_timestamps = [t for t in rpm_request_timestamps if t > now - 60]
                if len(rpm_request_timestamps) < rpm_limit:
                    rpm_request_timestamps.append(now)
                    break
            time.sleep(1)
    
    time.sleep(0.5) # Base throttle
    try:
        return model.generate_content(prompt).text.strip()
    except Exception as e:
        if args.rotate_on_limit_or_error:
            rotate_key(args)
            raise e # Retry trigger
        raise e

def translate_batch(entries, args, stop_event):
    """Logica Smart Batching"""
    # 1. Crea Batch Ibridi
    batches = []
    current_batch = []
    current_tokens = 0
    limit_tokens = 3000 # Max token per chiamata
    
    for entry in entries:
        text = entry['text']
        # Check Cache
        cache_key = json.dumps((text, args.source_lang, args.target_lang), ensure_ascii=False)
        if cache_key in translation_cache:
            entry['callback'](translation_cache[cache_key])
            continue

        toks = len(text) // ESTIMATED_CHARS_PER_TOKEN
        
        if (len(current_batch) >= args.batch_size) or (current_tokens + toks > limit_tokens):
            batches.append(current_batch)
            current_batch = []
            current_tokens = 0
        
        current_batch.append(entry)
        current_tokens += toks
    
    if current_batch: batches.append(current_batch)

    # 2. Processa Batch
    total_b = len(batches)
    for i, batch in enumerate(batches):
        if stop_event and stop_event.is_set(): return

        texts = [x['text'] for x in batch]
        log_msg(f"    ☁️  Batch {i+1}/{total_b} ({len(texts)} frasi)...", style="dim")

        prompt = f"""TRADUZIONE JSON ARRAY.
Da: {args.source_lang} | A: {args.target_lang}
Rispondi SOLO con un Array JSON di stringhe. Mantieni l'ordine.
Input:
{json.dumps(texts, ensure_ascii=False)}"""

        try:
            resp = call_ai_raw(prompt, args)
            clean = re.sub(r'^```json\s*|\s*```$', '', resp, flags=re.MULTILINE)
            trads = json.loads(clean)
            
            if len(trads) != len(texts): raise ValueError("Length mismatch")
            
            for idx, t in enumerate(trads):
                ck = json.dumps((texts[idx], args.source_lang, args.target_lang), ensure_ascii=False)
                translation_cache[ck] = t
                batch[idx]['callback'](t)
                
        except Exception as e:
            log_msg(f"⚠️ Batch fallito ({e}). Riprovo singolarmente.", style="yellow")
            # Fallback singolo
            for item in batch:
                try:
                    r = call_ai_raw(f"Traduci in {args.target_lang}: {item['text']}", args)
                    item['callback'](r)
                except: pass

# ----- HANDLERS FILE -----

def process_csv(fpath, outpath, args, stop_event):
    with open(fpath, 'r', encoding='utf-8', newline='') as f:
        rows = list(csv.reader(f, delimiter=args.delimiter))
    entries = []
    out_rows = [r[:] for r in rows]
    
    for i, row in enumerate(out_rows):
        if i > 0 and len(row) > args.translate_col: # Skip header approx
            txt = row[args.translate_col]
            if txt.strip() and not txt.isdigit():
                def cb(t, r=row, c=args.output_col):
                    while len(r) <= c: r.append('')
                    r[c] = t
                entries.append({'text': txt, 'callback': cb})
                
    translate_batch(entries, args, stop_event)
    with open(outpath, 'w', encoding='utf-8', newline='') as f:
        csv.writer(f, delimiter=args.delimiter).writerows(out_rows)

def process_json(fpath, outpath, args, stop_event):
    with open(fpath, 'r', encoding='utf-8') as f: data = json.load(f)
    entries = []
    keys = set(args.json_keys.split(',')) if args.json_keys else set()
    
    def traverse(obj, path=""):
        if isinstance(obj, dict):
            for k, v in obj.items():
                curr = f"{path}.{k}" if path else k
                match = curr in keys if args.match_full_json_path else k in keys
                if match and isinstance(v, str):
                    entries.append({'text': v, 'callback': lambda t, o=obj, key=k: o.__setitem__(key, t)})
                traverse(v, curr)
        elif isinstance(obj, list):
            for i, x in enumerate(obj): traverse(x, f"{path}[{i}]")
            
    traverse(data)
    translate_batch(entries, args, stop_event)
    with open(outpath, 'w', encoding='utf-8') as f: json.dump(data, f, indent=4, ensure_ascii=False)

def process_xlsx(fpath, outpath, args, stop_event):
    if not openpyxl:
        log_msg("❌ Modulo 'openpyxl' mancante. Impossibile processare Excel.", style="red")
        return
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    entries = []
    
    # Converte lettere (A, B) in indici 0-based se necessario, qui assumiamo input A/B
    # Semplificazione: Col A -> Col B
    for row in ws.iter_rows():
        cell = row[0] # Colonna A
        if cell.value and isinstance(cell.value, str):
            tgt_cell = ws.cell(row=cell.row, column=2) # Colonna B
            entries.append({'text': cell.value, 'callback': lambda t, c=tgt_cell: setattr(c, 'value', t)})
            
    translate_batch(entries, args, stop_event)
    wb.save(outpath)

# ----- ENTRY POINT PRINCIPALE -----

def run_core_process(args, log_queue=None, stop_event=None):
    """Funzione chiamata dalla GUI o dal main CLI."""
    global gui_log_queue, total_files_translated
    gui_log_queue = log_queue
    
    if not setup_engine(args): return

    files = [os.path.join(r, f) for r, _, fs in os.walk(args.input) for f in fs if f.lower().endswith(f".{args.file_type}")]
    if not files:
        log_msg(f"❌ Nessun file .{args.file_type} trovato in {args.input}", style="red")
        return

    base_out = f"{args.input}_tradotto"
    log_msg(f"🚀 Avvio Alumen 2.0 su {len(files)} file...", style="bold green")
    
    for i, fpath in enumerate(files):
        if stop_event and stop_event.is_set():
            log_msg("🛑 Processo interrotto dall'utente.", style="red")
            break
            
        fname = os.path.basename(fpath)
        log_msg(f"📄 [{i+1}/{len(files)}] Elaborazione: {fname}")
        
        rel = os.path.relpath(fpath, args.input)
        out = os.path.join(base_out, rel)
        os.makedirs(os.path.dirname(out), exist_ok=True)
        
        try:
            if args.file_type == 'csv': process_csv(fpath, out, args, stop_event)
            elif args.file_type == 'json': process_json(fpath, out, args, stop_event)
            elif args.file_type == 'xlsx': process_xlsx(fpath, out, args, stop_event)
            # Aggiungere PO/SRT qui se necessario con logica analoga
            
            total_files_translated += 1
            save_cache(args)
        except Exception as e:
            log_msg(f"❌ Errore critico su {fname}: {e}", style="bold red")

    log_msg(f"✅ Lavoro completato! {total_files_translated} file tradotti.", style="bold green")

# ----- CLI SUPPORT -----
def get_cli_args():
    parser = argparse.ArgumentParser(description="Alumen 2.0 CLI")
    parser.add_argument("--input", default="input")
    parser.add_argument("--api", help="API Key")
    parser.add_argument("--file-type", default="csv")
    parser.add_argument("--batch-size", type=int, default=30)
    parser.add_argument("--source-lang", default="inglese")
    parser.add_argument("--target-lang", default="italiano")
    parser.add_argument("--glossary")
    parser.add_argument("--persistent-cache", action="store_true")
    parser.add_argument("--rotate-on-limit-or-error", action="store_true")
    parser.add_argument("--rpm", type=int)
    parser.add_argument("--json-keys")
    parser.add_argument("--match-full-json-path", action="store_true")
    
    # Defaults nascosti
    args = parser.parse_args()
    args.model_name = DEFAULT_MODEL_NAME
    args.delimiter = ","
    args.translate_col = 3
    args.output_col = 3
    return args

if __name__ == "__main__":
    # Esecuzione diretta da terminale
    args = get_cli_args()
    run_core_process(args)
