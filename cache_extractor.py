import os
import json
import argparse
import csv
import polib
import re
from argparse_color_formatter import ColorHelpFormatter

# Tentativo import openpyxl per Excel
try:
    import openpyxl
except ImportError:
    openpyxl = None

# Nome del file di cache di default
CACHE_FILE_NAME = "alumen_cache.json"

# --- Funzioni di supporto ---

def log_critical_error_and_exit(message):
    print(f"🛑 ERRORE CRITICO: {message}")
    exit(1)

def determine_if_translatable(text_value):
    """Logica identica ad AlumenCore per determinare se una stringa va in cache."""
    if not isinstance(text_value, str): return False
    text_value_stripped = text_value.strip()
    if not text_value_stripped or text_value_stripped.isdigit() or re.match(r'^[\W_]+$', text_value_stripped) or "\\u" in text_value_stripped:
        return False
    # Filtra variabili isolate es. {name} o <br> (Allineato con AlumenCore v2.5)
    if re.match(r'^\{[\w\.]+\}$', text_value_stripped) or re.match(r'^<[\w\s="/]+>$', text_value_stripped): 
        return False
    return True

def get_cache_key(original_text, args):
    """
    Genera la chiave di cache.
    ⚠️ IMPORTANTE: Allineato con AlumenCore v2.0/2.5
    Formato: (text, source_lang, target_lang)
    """
    # Nota: game_name e context sono stati rimossi dalla chiave nella v2.0 per favorire il riutilizzo
    cache_key_tuple = (original_text, args.source_lang, args.target_lang)
    return json.dumps(cache_key_tuple, ensure_ascii=False)

# --- Logica di Estrazione ---

def extract_from_csv(s_path, t_path, cache, args):
    try:
        with open(s_path, 'r', encoding=args.encoding) as sf, open(t_path, 'r', encoding=args.encoding) as tf:
            s_rows = list(csv.reader(sf, delimiter=args.delimiter))
            t_rows = list(csv.reader(tf, delimiter=args.delimiter))
    except Exception as e: print(f"  ❌ Err CSV {s_path}: {e}"); return

    added = 0
    for i in range(min(len(s_rows), len(t_rows))):
        if i == 0 and not args.no_header: continue # Skip header
        s_r, t_r = s_rows[i], t_rows[i]
        if len(s_r) > args.source_col and len(t_r) > args.target_col:
            src, tgt = s_r[args.source_col], t_r[args.target_col]
            if determine_if_translatable(src) and tgt.strip():
                key = get_cache_key(src, args)
                if key not in cache:
                    cache[key] = tgt
                    added += 1
    print(f"  ✅ CSV: Aggiunte {added} voci.")

def extract_from_json(s_path, t_path, cache, args):
    try:
        with open(s_path, 'r', encoding=args.encoding) as sf, open(t_path, 'r', encoding=args.encoding) as tf:
            s_data, t_data = json.load(sf), json.load(tf)
    except Exception as e: print(f"  ❌ Err JSON {s_path}: {e}"); return

    # Fix: Gestione json_keys None
    if not args.json_keys:
        print(f"  ⚠️ Skip JSON {s_path}: Nessuna chiave specificata (--json-keys).")
        return

    keys = set(args.json_keys.split(','))
    added = 0
    
    def traverse(s_obj, t_obj, path=""):
        nonlocal added
        if isinstance(s_obj, dict) and isinstance(t_obj, dict):
            for k, v in s_obj.items():
                curr = f"{path}.{k}" if path else k
                is_match = (curr in keys) if args.match_full_json_path else (k in keys)
                tgt_v = t_obj.get(k)
                if is_match and isinstance(v, str) and isinstance(tgt_v, str) and determine_if_translatable(v) and tgt_v.strip():
                    key = get_cache_key(v, args)
                    if key not in cache:
                        cache[key] = tgt_v
                        added += 1
                if k in t_obj: traverse(v, t_obj.get(k), curr)
        elif isinstance(s_obj, list) and isinstance(t_obj, list):
            for i in range(min(len(s_obj), len(t_obj))): traverse(s_obj[i], t_obj[i], path)

    traverse(s_data, t_data)
    print(f"  ✅ JSON: Aggiunte {added} voci.")

def extract_from_po(t_path, cache, args):
    try:
        po = polib.pofile(t_path, encoding=args.encoding)
    except Exception as e: print(f"  ❌ Err PO {t_path}: {e}"); return
    added = 0
    for e in po:
        if e.msgid and e.msgstr and determine_if_translatable(e.msgid):
            key = get_cache_key(e.msgid, args)
            if key not in cache:
                cache[key] = e.msgstr
                added += 1
    print(f"  ✅ PO: Aggiunte {added} voci.")

def extract_from_srt(s_path, t_path, cache, args):
    try:
        with open(s_path, 'r', encoding=args.encoding) as f: s_txt = f.read()
        with open(t_path, 'r', encoding=args.encoding) as f: t_txt = f.read()
    except Exception: return

    # Regex SRT semplice
    pattern = re.compile(r'\d+\s*\n\d{2}:\d{2}:\d{2},\d{3}\s*-->\s*\d{2}:\d{2}:\d{2},\d{3}\s*\n(.*?)(?=\n\s*\n|\Z)', re.DOTALL)
    s_matches = [m.group(1).strip() for m in pattern.finditer(s_txt)]
    t_matches = [m.group(1).strip() for m in pattern.finditer(t_txt)]

    added = 0
    for src, tgt in zip(s_matches, t_matches):
        if determine_if_translatable(src) and tgt:
            key = get_cache_key(src, args)
            if key not in cache:
                cache[key] = tgt
                added += 1
    print(f"  ✅ SRT: Aggiunte {added} voci.")

def extract_from_xlsx(s_path, t_path, cache, args):
    if not openpyxl: print("  ⚠️ Modulo openpyxl mancante. Salto XLSX."); return
    try:
        wb_s = openpyxl.load_workbook(s_path, read_only=True)
        wb_t = openpyxl.load_workbook(t_path, read_only=True)
        ws_s = wb_s.active
        ws_t = wb_t.active
    except Exception as e: print(f"  ❌ Err XLSX {s_path}: {e}"); return

    # Converti lettere colonna (A, B) in indici 0-based
    try:
        s_cols_str = [c.strip() for c in str(args.xlsx_source_col).split(',') if c.strip()]
        t_cols_str = [c.strip() for c in str(args.xlsx_target_col).split(',') if c.strip()]
        if len(t_cols_str) < len(s_cols_str): t_cols_str.extend([t_cols_str[-1]] * (len(s_cols_str) - len(t_cols_str)))
        s_indices = [openpyxl.utils.column_index_from_string(c) - 1 for c in s_cols_str]
        t_indices = [openpyxl.utils.column_index_from_string(c) - 1 for c in t_cols_str]
    except: print("  ❌ Colonne Excel non valide."); return

    added = 0
    # Itera sulle righe (zip per sicurezza)
    for r_s, r_t in zip(ws_s.iter_rows(), ws_t.iter_rows()):
        for sc, tc in zip(s_indices, t_indices):
            if len(r_s) > sc and len(r_t) > tc:
                src = r_s[sc].value
                tgt = r_t[tc].value
                if isinstance(src, str) and isinstance(tgt, str) and determine_if_translatable(src) and tgt.strip():
                    key = get_cache_key(src, args)
                    if key not in cache:
                        cache[key] = tgt
                        added += 1
    print(f"  ✅ XLSX: Aggiunte {added} voci.")

# --- Main ---

def process_files(args, cache):
    s_dir, t_dir = os.path.abspath(args.source_dir), os.path.abspath(args.target_dir)
    print(f"Scansione: {s_dir} -> {t_dir} (Format: {args.file_type})")
    
    for root, _, files in os.walk(s_dir):
        rel = os.path.relpath(root, s_dir)
        curr_t_dir = os.path.join(t_dir, rel)
        target_files = [f for f in files if f.endswith(f'.{args.file_type}')]
        
        for fname in target_files:
            s_path = os.path.join(root, fname)
            t_path = os.path.join(curr_t_dir, fname)
            
            if args.file_type == 'po': # PO file has both src and tgt
                t_path = s_path 
            
            if not os.path.exists(t_path) and args.file_type != 'po':
                continue

            print(f" Elaborazione: {fname}")
            if args.file_type == 'csv': extract_from_csv(s_path, t_path, cache, args)
            elif args.file_type == 'json': extract_from_json(s_path, t_path, cache, args)
            elif args.file_type == 'po': extract_from_po(s_path, cache, args)
            elif args.file_type == 'srt': extract_from_srt(s_path, t_path, cache, args)
            elif args.file_type == 'xlsx': extract_from_xlsx(s_path, t_path, cache, args)

def main():
    parser = argparse.ArgumentParser(description="Alumen Cache Extractor v2.6.2", formatter_class=ColorHelpFormatter)
    
    g = parser.add_argument_group('Configurazione')
    g.add_argument("--source-dir", required=True, help="Cartella file ORIGINALI")
    g.add_argument("--target-dir", required=True, help="Cartella file TRADOTTI")
    g.add_argument("--file-type", default="csv", choices=['csv', 'json', 'po', 'srt', 'xlsx'])
    g.add_argument("--encoding", default="utf-8")
    g.add_argument("--output-cache-file", default=CACHE_FILE_NAME)
    g.add_argument("--append", action="store_true")

    c = parser.add_argument_group('CSV/Excel')
    c.add_argument("--delimiter", default=",")
    c.add_argument("--source-col", type=int, default=3, help="Index colonna originale (CSV)")
    c.add_argument("--target-col", type=int, default=3, help="Index colonna tradotta (CSV)")
    c.add_argument("--xlsx-source-col", type=str, default="A", help="Lettera colonna/e originale Excel (es. A,C)")
    c.add_argument("--xlsx-target-col", type=str, default="B", help="Lettera colonna/e tradotta Excel")
    c.add_argument("--no-header", action="store_true")

    j = parser.add_argument_group('JSON')
    j.add_argument("--json-keys", default=None)
    j.add_argument("--match-full-json-path", action="store_true")

    t = parser.add_argument_group('Parametri Traduzione (Devono coincidere con AlumenCore)')
    t.add_argument("--source-lang", default="inglese")
    t.add_argument("--target-lang", default="italiano")
    # Nota: game-name e prompt-context rimossi dalla chiave cache v2.0 per universalità

    args = parser.parse_args()
    if args.delimiter == '\\t': args.delimiter = '\t'

    cache = {}
    if args.append and os.path.exists(args.output_cache_file):
        try: 
            with open(args.output_cache_file, 'r', encoding='utf-8') as f: cache = json.load(f)
            print(f"Cache caricata: {len(cache)} voci.")
        except: pass

    process_files(args, cache)

    if cache:
        with open(args.output_cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=4)
        print(f"🎉 Cache salvata: {args.output_cache_file} ({len(cache)} voci).")
    else:
        print("Nessuna voce estratta.")

if __name__ == "__main__":
    main()
